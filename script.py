import os
import re
import time
import evdev
import pyudev
import logging
import smtplib
import schedule
import threading
import configparser
from datetime import datetime
from collections import deque
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pyModbusTCP.client import ModbusClient
from openpyxl import Workbook, load_workbook

# Configure logging
logging.basicConfig(filename='app.log', 
                    level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

logging.info("Program started.")

# Initialize an empty queue
queue = deque()

def enqueue(item):
    """Add an item to the end of the queue."""
    logging.info(f"Enqueuing item: {item}")
    queue.append(item)

def dequeue():
    """Remove and return an item from the front of the queue. Raises an exception if the queue is empty."""
    if is_empty():
        logging.error("Attempt to dequeue from an empty queue.")
        raise IndexError("Dequeue from an empty queue")
    item = queue.popleft()
    logging.info(f"Dequeued item: {item}")
    return item

def peek():
    """Return the item at the front of the queue without removing it. Raises an exception if the queue is empty."""
    if is_empty():
        logging.error("Attempt to peek from an empty queue.")
        raise IndexError("Peek from an empty queue")
    item = queue[0]
    logging.info(f"Peeked at item: {item}")
    return item

def is_empty():
    """Check if the queue is empty."""
    return len(queue) == 0

def size():
    """Return the number of items in the queue."""
    return len(queue)

# Modbus TCP client setup
modbus_client = ModbusClient(host='192.168.1.7', port=502)

scanning_mode = None
scan_started = False
pallet_data = None
reels_data = []
last_reels_data = []
count = 0


def read_credentials():
    config_file = 'config.ini'
    config = configparser.ConfigParser()
    config.read(config_file)

    sender_email = config.get('credentials', 'sender_email')
    receiver_email = config.get('credentials', 'receiver_email').split(',')  # Split by comma
    password = config.get('credentials', 'app_password')
    smtp_server = config.get('credentials', 'smtp_server')
    smtp_port = config.get('credentials', 'smtp_port')

    return sender_email, receiver_email, password, smtp_server, smtp_port

def send_email(subject, body, file_path=None):
    sender_email, receiver_emails, password, smtp_server, smtp_port = read_credentials()

    # Create the SMTP server connection once
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()  # Upgrade the connection to TLS
        server.login(sender_email, password)  # Login to SMTP server

        for receiver_email in receiver_emails:
            # Create a new email message for each recipient
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = receiver_email.strip()  # Set the To header
            msg['Subject'] = subject

            # Attach the email body
            msg.attach(MIMEText(body, 'plain'))

            # Attach file if file_path is provided
            if file_path and os.path.exists(file_path):
                filename = os.path.basename(file_path)
                with open(file_path, 'rb') as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', f'attachment; filename={filename}')
                    msg.attach(part)
                    logging.info(f"Attached file: {filename}")

            try:
                server.sendmail(sender_email, receiver_email.strip(), msg.as_string())
                logging.info(f"Email sent successfully to {receiver_email.strip()}")
            except smtplib.SMTPException as e:
                logging.error(f"Failed to send email to {receiver_email.strip()}: {e}")


def format_reel_data_email(reel_data):
    body = "Reel Data Summary:\n\n"
    
    # Add timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    body += f"Timestamp: {timestamp}\n\n"

    # Identify mismatches in production orders
    first_order = reel_data[0]['production_order']
    mismatches = [obj for obj in reel_data if obj['production_order'] != first_order]

    # Add summary of mismatches
    if mismatches:
        body += f"Total Mismatches: {len(mismatches)} found out of {len(reel_data)} reels.\n\n"
    else:
        body += "No mismatches detected.\n\n"

    # Add details for each reel entry
    for idx, reel in enumerate(reel_data):
        body += (
            f"Reel {idx + 1}:\n"
            f"Production Order: {reel['production_order']}\n"
            f"Reel Number: {reel['reel_number']}\n"
            f"Var Count: {reel['var_count']}\n\n"
        )

    # Highlight mismatches
    if mismatches:
        body += "Mismatched Production Orders:\n\n"
        for mismatch in mismatches:
            body += (
                f"Production Order: {mismatch['production_order']} (Mismatched)\n"
                f"Reel Number: {mismatch['reel_number']}\n"
                f"Var Count: {mismatch['var_count']}\n\n"
            )

    # Add closing
    body += "Please review the above data and take the necessary actions.\n"
    body += "\nBest regards"

    # Create additional row data
    date_time = timestamp
    production_order_reels = first_order
    reel_numbers = ', '.join(reel['reel_number'] for reel in reel_data)  # Comma-separated
    var_counts = ', '.join(str(reel['var_count']) for reel in reel_data)  # Comma-separated
    production_order_pallet = None
    pallet_contents = None
    status = "Fail"
    station = 1

    row = [
        date_time,
        production_order_reels,
        reel_numbers,
        var_counts,
        production_order_pallet,
        pallet_contents,
        status,
        station
    ]
    
    return body, row

def generate_mismatch_email(reel_data, pallet_data):
    email_body = "Dear Team,\n\n"
    email_body += "A mismatch has been detected between the reel data and pallet data. Please review the details below:\n\n"

    # Extract production orders
    production_order_reel = reel_data[0]['production_order'] if reel_data else "N/A"
    production_order_pallet = pallet_data['production_order']

    # Check if production orders match
    if production_order_reel != production_order_pallet:
        email_body += f"*** Production Order Mismatch! ***\n"
        email_body += f"Production Order from Reel Data: {production_order_reel}\n"
        email_body += f"Production Order from Pallet Data: {production_order_pallet}\n\n"
    else:
        email_body += f"Production Order: {production_order_pallet} (Matches in both Reel and Pallet)\n\n"

    # Reel Data Details
    email_body += "Reel Data:\n"
    for idx, reel in enumerate(reel_data, start=1):
        email_body += f"Reel {idx}:\n"
        email_body += f"- Reel Number: {reel['reel_number']}\n"
        email_body += f"- Var Count: {reel['var_count']}\n\n"

    # Pallet Data Details
    email_body += "Pallet Data:\n"
    email_body += f"Production Order: {pallet_data['production_order']}\n"
    email_body += "Pallet Contents:\n"
    for item in pallet_data['pallet_contents']:
        email_body += f"- {item}\n"

    email_body += "\n\nThe reel and pallet data mismatch requires immediate attention. Please resolve the issue.\n\n"
    email_body += "Best regards"

    return email_body

def read_coil(address):
    """ Read the state of a coil. """
    result = modbus_client.read_coils(address, 1)
    if result is None:
        logging.error(f"Failed to read coil at address {address}.")
        return False
    return result[0]

def write_coil(address, value):
    """ Write a value to a coil. """
    if not modbus_client.write_single_coil(address, value):
        logging.error(f"Failed to write to coil at address {address}.")
    else:
        logging.info(f"Coil at address {address} set to {value}")

def write_register(address, value):
    """ Write a value to a register. """
    if not modbus_client.write_single_register(address, value):
        logging.error(f"Failed to write to register at address {address}.")
    else:
        logging.info(f"Register at address {address} set to {value}")

def verify_production_orders(objects):
    if not objects:
        return True  # If the list is empty, consider it true as there are no differing orders

    # Extract production orders from objects
    production_orders = [obj['production_order'] for obj in objects]
    
    # Check if all production orders are the same
    first_order = production_orders[0]
    return all(order == first_order for order in production_orders)

def connect_to_plc():
    """ Connect to the PLC. """
    if not modbus_client.open():
        logging.error("Failed to connect to PLC.")
        return False
    logging.info("Connected to PLC successfully.")
    return True

def plc_communication():
    logging.info("Starting PLC communication thread.")

    global scan_started, scanning_mode, pallet_data, reels_data, last_reels_data, count

    # Attempt to connect to PLC in a loop until successful
    while not connect_to_plc():
        logging.warning("Retrying connection to PLC in 5 seconds...")
        time.sleep(5)

    while True:
        try:
            if read_coil(8):  # Address for coil m8
                if not scan_started:
                    logging.info("m8 is TRUE: Processing PLC logic (Start Scan Reels)")
                    scan_started = True
                    scanning_mode = "reel"
                    reels_data = []
                    count = 0
                    write_register(10, count)

            if read_coil(12):  # Address for coil m12
                logging.info("m12 is TRUE: Processing PLC logic (Scan complete)")
                write_coil(12, False)  # Reset Coil 12

                if reels_data:
                    success = verify_production_orders(reels_data)

                    if success:
                        logging.info("Production orders correct!")
                        write_coil(14, True)  # Write to coil m14
                        # enqueue(reels_data)
                        last_reels_data = reels_data
                    else:
                        logging.warning("Production orders incorrect!")
                        write_coil(16, True)  # Write to coil m16

                        subject = 'Reel Data Mismatch Notification'
                        body, row_data = format_reel_data_email(reels_data)
                        append_row(row_data)
                        append_row(row_data, report_type='monthly')
                        # send_email(subject, body)

                # Reset Global Variables
                scan_started = False
                scanning_mode = None
                reels_data = []
                count = 0

            if read_coil(40):  # Address for coil m40
                if not scan_started:
                    logging.info("m40 is TRUE: Processing PLC logic (Start Scan Pallet)")
                    scan_started = True
                    scanning_mode = "pallet"

        except Exception as e:
            logging.error(f"Error during PLC communication: {e}")
            logging.warning("Attempting to reconnect to PLC in 5 seconds...")
            time.sleep(5)  # Wait before retrying connection
            # Try to reconnect
            while not connect_to_plc():
                logging.warning("Retrying connection to PLC in 5 seconds...")
                time.sleep(5)

        # Sleep to prevent high CPU usage
        time.sleep(1)

def verify_data(reel_data, pallet_data):
    # Extract production order and contents from the input data
    production_order = pallet_data['production_order']
    pallet_contents = pallet_data['pallet_contents']
    
    # Convert pallet contents into a set of tuples (reel_number, var_count)
    pallet_contents_set = set()
    for item in pallet_contents:
        reel_number, var_count = item.split(' / ')
        pallet_contents_set.add((reel_number, var_count))
    
    # Extract production data from the list of dictionaries
    reel_data_set = set()
    for item in reel_data:
        if item['production_order'] == production_order:
            reel_data_set.add((item['reel_number'], item['var_count']))
    
    # Check if both sets match
    return pallet_contents_set == reel_data_set

def extract_data(s):
    marker = 'FNC103'
    first_dash = s.find('-')
    second_dash = s.find('-', first_dash + 1)
    if second_dash == -1:
        return {"message": "Second dash not found", "success": False}
    start_of_interest = second_dash + 5
    marker_index = s.find(marker)
    if marker_index == -1:
        return {"message": "Marker not found", "success": False}
    count = s[start_of_interest:marker_index]
    production_order_start = marker_index + len(marker)
    raw_production_order = s[production_order_start:production_order_start + 11]
    production_order = f"P{raw_production_order[:3]}-{raw_production_order[4:]}"
    reel_number_start = production_order_start + 11
    raw_reel_number = s[reel_number_start:reel_number_start + 6]
    reel_number = f"{int(raw_reel_number[:2])}-{raw_reel_number[2:]}"
    return {'var_count': count, 'production_order': production_order, 'reel_number': reel_number, "success": True}

def extract_pallet_contents(input_string):
    # Define regex pattern to verify pallet format
    pallet_pattern = r'^\d+-\d+ \/ \d+$'

    # Find the first dash and attempt to get first_pallet_content
    first_dash_index = input_string.find('-')
    first_pallet_content_start = first_dash_index - 1
    first_pallet_content_end = input_string.find(',', first_dash_index)
    first_pallet_content = input_string[first_pallet_content_start:first_pallet_content_end]

    # If first_pallet_content doesn't match the pallet pattern, find the last dash and re-calculate
    if not re.match(pallet_pattern, first_pallet_content):
        last_dash_index = first_pallet_content.rfind('-')
        if last_dash_index != -1:
            first_pallet_content_start = last_dash_index - 1
            first_pallet_content = first_pallet_content[first_pallet_content_start:]

    # Split the rest of the contents
    remaining_pallet_contents = input_string[first_pallet_content_end + 1:].split(',')
    pallet_contents = [first_pallet_content] + remaining_pallet_contents
    pallet_contents = [p.strip() for p in pallet_contents if p.strip()]

    # Modify pallet contents to convert numbers after the slash
    modified_pallet_contents = []
    for content in pallet_contents:
        if ' / ' in content:
            part_before_slash, part_after_slash = content.split(' / ')
            part_after_slash_int = str(int(part_after_slash))  # Remove leading zeros by converting to int
            modified_pallet_contents.append(f"{part_before_slash} / {part_after_slash_int}")
        else:
            modified_pallet_contents.append(content)

    # Extract and format production order
    first_10_chars = input_string[:10]
    modified_chars = first_10_chars[3:]
    
    return {
        "pallet_contents": modified_pallet_contents,
        "production_order": "P552-" + modified_chars,
        "success": True
    }

def find_device(vendor_id, product_id):
    context = pyudev.Context()
    for device in context.list_devices(subsystem='usb'):
        if vendor_id in device.get('ID_VENDOR_ID', '') and product_id in device.get('ID_MODEL_ID', ''):
            for child in device.children:
                if child.device_node and 'event' in child.device_node:
                    return child.device_node
    return None

vendor_id = '8888'
product_id = '2019'

key_map = {
    'KEY_1': '1', 'KEY_2': '2', 'KEY_3': '3', 'KEY_4': '4', 'KEY_5': '5',
    'KEY_6': '6', 'KEY_7': '7', 'KEY_8': '8', 'KEY_9': '9', 'KEY_0': '0',
    'KEY_A': 'A', 'KEY_B': 'B', 'KEY_C': 'C', 'KEY_D': 'D', 'KEY_E': 'E',
    'KEY_F': 'F', 'KEY_G': 'G', 'KEY_H': 'H', 'KEY_I': 'I', 'KEY_J': 'J',
    'KEY_K': 'K', 'KEY_L': 'L', 'KEY_M': 'M', 'KEY_N': 'N', 'KEY_O': 'O',
    'KEY_P': 'P', 'KEY_Q': 'Q', 'KEY_R': 'R', 'KEY_S': 'S', 'KEY_T': 'T',
    'KEY_U': 'U', 'KEY_V': 'V', 'KEY_W': 'W', 'KEY_X': 'X', 'KEY_Y': 'Y',
    'KEY_Z': 'Z',
    'KEY_F1': 'F1', 'KEY_F2': 'F2', 'KEY_F3': 'F3', 'KEY_F4': 'F4',
    'KEY_F5': 'F5', 'KEY_F6': 'F6', 'KEY_F7': 'F7', 'KEY_F8': 'F8',
    'KEY_F9': 'F9', 'KEY_F10': 'F10', 'KEY_F11': 'F11', 'KEY_F12': 'F12',
    'KEY_LEFT': '<', 'KEY_RIGHT': '>', 'KEY_UP': '^', 'KEY_DOWN': 'v',
    'KEY_ENTER': '\n', 'KEY_SPACE': ' ', 'KEY_TAB': '\t',
    'KEY_MINUS': '-', 'KEY_EQUAL': '=', 'KEY_LEFTBRACE': '[', 'KEY_RIGHTBRACE': ']',
    'KEY_BACKSLASH': '\\', 'KEY_SEMICOLON': ';', 'KEY_APOSTROPHE': "'",
    'KEY_GRAVE': '`', 'KEY_COMMA': ',', 'KEY_DOT': '.', 'KEY_SLASH': '/',
    'KEY_CAPSLOCK': 'CAPSLOCK', 'KEY_NUMLOCK': '',  # Ignore NUMLOCK
    'KEY_SCROLLLOCK': 'SCROLLLOCK',
    'KEY_ESC': 'ESC', 'KEY_BACKSPACE': 'BACKSPACE', 'KEY_INSERT': 'INSERT',
    'KEY_DELETE': 'DELETE', 'KEY_HOME': 'HOME', 'KEY_END': 'END', 'KEY_PAGEUP': 'PAGEUP',
    'KEY_PAGEDOWN': 'PAGEDOWN', 'KEY_UP': 'UP', 'KEY_DOWN': 'DOWN', 'KEY_LEFT': 'LEFT',
    'KEY_RIGHT': 'RIGHT', 'KEY_PRINT': 'PRINT', 'KEY_PAUSE': 'PAUSE',
    'KEY_LCTRL': 'LCTRL', 'KEY_LSHIFT': 'LSHIFT', 'KEY_LALT': 'LALT', 'KEY_LGUI': 'LGUI',
    'KEY_RCTRL': 'RCTRL', 'KEY_RSHIFT': 'RSHIFT', 'KEY_RALT': 'RALT', 'KEY_RGUI': 'RGUI',
    'KEY_APPLICATION': 'APPLICATION', 'KEY_POWER': 'POWER',
    'KEY_HELP': 'HELP', 'KEY_MENU': 'MENU', 'KEY_SELECT': 'SELECT',
}

def wait_for_device(vendor_id, product_id):
    run_once = False
    while True:
        device_path = find_device(vendor_id, product_id)
        if device_path:
            logging.info(f"Device found: {device_path}")
            write_coil(55, False)  # Write to coil m14
            return device_path
        else:
            logging.info("Device not found. Waiting for reconnect...")
            write_coil(55, True)  # Write to coil m14
            # if not run_once:
            #     logging.info("Device not found. Waiting for reconnect...")
            #     write_coil(55, True)  # Write to coil m14
            #     run_once = True
            time.sleep(1)  # Wait before trying again

def log_scanner(message):
    with open('scan_data.log', 'a') as f:
        current_time = time.strftime('%Y-%m-%d %H:%M:%S')  # Get the current time
        f.write(f'{current_time} - {message}\n')  # Write the message with timestamp

def add_reel_data(scan_data):
    global reels_data

    """Add reel data if it's not already in the list and update count."""
    if scan_data not in reels_data:
        reels_data.append(scan_data)
        logging.info(f"Added reel data: {scan_data}")
        return True  # Indicate that data was added
    else:
        logging.info(f"Duplicate reel data not added: {scan_data}")
        return False  # Indicate that data was not added

def get_file_path(report_type='daily'):
    """
    Get the file path for daily or monthly reports.
    
    :param report_type: 'daily' (default) or 'monthly' to specify the type of report
    :return: The full file path for the Excel report
    """
    # Get current working directory
    current_dir = os.getcwd()
    
    # Define the reports folder
    reports_folder = os.path.join(current_dir, 'reports')
    
    # For monthly reports, create a 'monthly' subfolder
    if report_type == 'monthly':
        reports_folder = os.path.join(reports_folder, 'monthly')
    
    # Create the reports folder if it doesn't exist
    if not os.path.exists(reports_folder):
        os.makedirs(reports_folder)
    
    # Generate the filename based on the report type
    if report_type == 'daily':
        today = datetime.now().strftime('%Y-%m-%d')
        file_name = f'{today}.xlsx'
    elif report_type == 'monthly':
        current_month = datetime.now().strftime('%Y-%m')
        file_name = f'{current_month}.xlsx'
    else:
        raise ValueError("Invalid report_type. Use 'daily' or 'monthly'.")
    
    # Full path to the Excel file
    file_path = os.path.join(reports_folder, file_name)
    
    return file_path

def create_excel_file(report_type='daily'):
    """
    Create an Excel file for daily or monthly reports if it doesn't exist.
    
    :param report_type: 'daily' (default) or 'monthly' to specify the type of report
    """
    file_path = get_file_path(report_type)

    if not os.path.exists(file_path):
        # Create a new workbook and add column headers
        wb = Workbook()
        ws = wb.active
        ws.append(['Date Time', 'Production Order Reels', 'Reels Numbers', 'Var Counts', 'Production Order Pallet', 'Pallet Contents', 'Status', 'Station'])  # Add your desired column names
        wb.save(file_path)
        logging.info(f"Excel file '{file_path}' created with headers.")
    else:
        logging.info(f"Excel file '{file_path}' already exists.")

def append_row(data, report_type='daily'):
    """
    Append a row to the Excel file for daily or monthly reports.
    
    :param data: The row data to append
    :param report_type: 'daily' (default) or 'monthly' to specify the type of report
    """
    file_path = get_file_path(report_type)

    # Check if the file exists, if not, create it
    if not os.path.exists(file_path):
        logging.warning(f"File '{file_path}' not found. Creating a new one.")
        create_excel_file(report_type)

    # Open the existing or newly created Excel file and select the active sheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Append the new row
    ws.append(data)
    
    # Save the workbook
    wb.save(file_path)
    logging.info(f"Row {data} appended to the {report_type} Excel file.")

def generate_excel_row(reel_data, pallet_data, success):
    # Get the current date and time
    date_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Extract the production order from the reels (assuming all reels have the same production order)
    production_order_reels = reel_data[0]['production_order']
    
    # Extract reel numbers and var counts as comma-separated strings
    reel_numbers = ', '.join([item['reel_number'] for item in reel_data])
    var_counts = ', '.join([item['var_count'] for item in reel_data])
    
    # Extract the production order and pallet contents
    production_order_pallet = pallet_data['production_order']
    pallet_contents = ', '.join(pallet_data['pallet_contents'])
    
    # Status based on the passed `success` parameter
    status = 'Success' if success else 'Fail'
    station = 5
    
    # Create the row data
    row = [
        date_time,
        production_order_reels,
        reel_numbers,
        var_counts,
        production_order_pallet,
        pallet_contents,
        status,
        station
    ]
    
    return row

def barcode_scanning():
    logging.info("Starting Barcode Scanning Thread.")

    global scan_started, scanning_mode, pallet_data, reels_data, last_reels_data, count
    
    while True:
        device_path = wait_for_device(vendor_id, product_id)
        device = evdev.InputDevice(device_path)
        logging.info(f"Opened device {device.name} at {device_path}")

        try:
            device.grab()
            logging.info("Device grabbed successfully.")
        except IOError as e:
            logging.error(f"Error grabbing device: {e}")

        barcode = ""

        try:
            for event in device.read_loop():
                if event.type == evdev.ecodes.EV_KEY:
                    key_event = evdev.categorize(event)
                    if key_event.keystate == key_event.key_down:
                        key = key_event.keycode
                        if key in key_map:
                            if key == 'KEY_ENTER':
                                # logging.info(f"Scanned barcode: {barcode}")
                                log_scanner(barcode)
                                if scanning_mode == "reel":
                                    scan_data = extract_data(barcode)
                                    if scan_data['success']:
                                        if add_reel_data(scan_data):
                                            count += 1
                                            write_register(10, count)

                                elif scanning_mode == "pallet":
                                    scan_data = extract_data(barcode)
                                    if not scan_data['success']:
                                        pallet_data = extract_pallet_contents(barcode)
                                        logging.info(f"Pallet data extracted: {pallet_data}")
                
                                        # reels_data = dequeue()
                                        reels_data = last_reels_data
                                        success = verify_data(reels_data, pallet_data)
                                        row_data = generate_excel_row(reels_data, pallet_data, success)
                                        append_row(row_data)
                                        append_row(row_data, report_type='monthly')
                                        if success: 
                                            logging.info("Data match correct!")
                                            write_coil(42, True)  # Write to coil m14
                                        else:
                                            logging.warning("Data match incorrect!")
                                            write_coil(44, True)  # Write to coil m16

                                            subject = "Reel and Pallet Data Mismatch Detected"
                                            body = generate_mismatch_email(reels_data, pallet_data)
                                            # send_email(subject, body)
                                        # if size() > 0:
                                        #     reels_data = dequeue()
                                        #     success = verify_data(reels_data, pallet_data)
                                        #     row_data = generate_excel_row(reels_data, pallet_data, success)
                                        #     append_row(row_data)
                                        #     if success: 
                                        #         logging.info("Data match correct!")
                                        #         write_coil(42, True)  # Write to coil m14
                                        #     else:
                                        #         logging.warning("Data match incorrect!")
                                        #         write_coil(44, True)  # Write to coil m16

                                        #         subject = "Reel and Pallet Data Mismatch Detected"
                                        #         body = generate_mismatch_email(reels_data, pallet_data)
                                        #         send_email(subject, body)
                                        # else:
                                        #     logging.warning("Queue is Empty!")
                                        #     write_coil(44, True)  # Write to coil m16

                                        # Reset Global Variables
                                        scan_started = False
                                        scanning_mode = None
                                        reels_data = []
                                        count = 0

                                barcode = ""
                            else:
                                barcode += key_map[key]
        except OSError:
            logging.error("Device removed or error occurred. Waiting for reconnect...")


def send_daily_report():
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    logging.log(f"Running task at {current_time}...")

    current_date = datetime.now().strftime('%Y-%m-%d')

    subject = "Daily Report"
    body = "Please find the attached report!"

    # # Sending email with attachment
    file_path = get_file_path()
    # send_email(subject, body, file_path=file_path)

    # Here you can add the code you want to run daily
    # For example: sending an email, creating a report, etc.
    logging.log("Daily task executed successfully!")

# Schedule the task to run every day at 23:59
schedule.every().day.at("23:59").do(send_daily_report)

# Function to keep the script running and checking the schedule
def run_scheduler():
    while True:
        schedule.run_pending()  # Run any scheduled task that is pending
        time.sleep(1)  # Sleep for a second to prevent CPU overuse

if __name__ == "__main__":
    # Start PLC thread
    plc_thread = threading.Thread(target=plc_communication)
    plc_thread.start()

    # Start barcode scanning thread
    barcode_thread = threading.Thread(target=barcode_scanning)
    barcode_thread.start()

    scheduler_thread = threading.Thread(target=run_scheduler)
    scheduler_thread.start()

    # Wait for both threads to complete
    plc_thread.join()
    barcode_thread.join()
    scheduler_thread.join()