import os
import time
from datetime import datetime
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk
from threading import Thread

# Function to get the current file path
def get_file_path():
    current_dir = os.getcwd()
    reports_folder = os.path.join(current_dir, 'reports')

    # Create the reports folder if it doesn't exist
    if not os.path.exists(reports_folder):
        os.makedirs(reports_folder)

    today = datetime.now().strftime('%Y-%m-%d')
    file_name = f'{today}.xlsx'
    file_path = os.path.join(reports_folder, file_name)

    return file_path

# Function to periodically read from the Excel file
def read_excel_data():
    while True:
        try:
            file_path = get_file_path()  # Get the file path dynamically

            # Load the workbook
            workbook = load_workbook(file_path, read_only=True)

            # Assume the sheet is the first one
            sheet = workbook.active

            # Clear the data in the treeview
            for row in tree.get_children():
                tree.delete(row)

            # Read data from Excel and update the UI
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                tree.insert("", "end", values=(row_index,) + row)  # Add Sr.

            workbook.close()  # Always close the workbook to avoid locking it
        except FileNotFoundError:
            print(f"File {file_path} not found!")
        except Exception as e:
            print(f"Error reading file: {e}")

        # Pause for 30 seconds before reading the file again
        time.sleep(30)

# Function to start the Excel reading thread
def start_excel_thread():
    excel_thread = Thread(target=read_excel_data)
    excel_thread.daemon = True  # This will stop the thread when the main program exits
    excel_thread.start()

# Function to toggle full screen mode
def toggle_fullscreen(event=None):
    global fullscreen
    fullscreen = not fullscreen
    root.attributes("-fullscreen", fullscreen)

# Function to exit full screen mode
def exit_fullscreen(event=None):
    global fullscreen
    fullscreen = False
    root.attributes("-fullscreen", False)

# Create a simple UI to display data
root = tk.Tk()
root.title("Reels Data Viewer")

# Set the global fullscreen variable
fullscreen = True
toggle_fullscreen()  # Start in full screen

# Include an additional column for Sr.
tree = ttk.Treeview(root, columns=('Sr.', 'Date Time', 'Production Order Reels', 'Reels Numbers', 'Var Counts', 'Production Order Pallet', 'Pallet Contents', 'Status', 'Station'), show='headings')

# Define headings and set width for 'Sr.' column
tree.heading('Sr.', text='Sr.')
tree.column('Sr.', width=30)

# Define widths for other columns
tree.heading('Date Time', text='Date Time')
tree.column('Date Time', width=120)
tree.heading('Production Order Reels', text='Production Order Reels')
tree.column('Production Order Reels', width=150)
tree.heading('Reels Numbers', text='Reels Numbers')
tree.column('Reels Numbers', width=120)
tree.heading('Var Counts', text='Var Counts')
tree.column('Var Counts', width=100)
tree.heading('Production Order Pallet', text='Production Order Pallet')
tree.column('Production Order Pallet', width=150)
tree.heading('Pallet Contents', text='Pallet Contents')
tree.column('Pallet Contents', width=150)
tree.heading('Status', text='Status')
tree.column('Status', width=100)
tree.heading('Station', text='Station')
tree.column('Station', width=100)

# Arrange the treeview in the window
tree.pack(fill=tk.BOTH, expand=True)

# Bind key events for toggling full screen
root.bind("<F11>", toggle_fullscreen)  # Press F11 to toggle full screen
root.bind("<Escape>", exit_fullscreen)  # Press Esc to exit full screen

# Start the background thread for reading Excel data
start_excel_thread()

# Start the Tkinter main loop (this is the UI loop)
root.mainloop()
